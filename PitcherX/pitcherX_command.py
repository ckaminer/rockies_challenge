from openpyxl import load_workbook
import plotly
import plotly.graph_objs as go
import sys
import os

# load data for pitcher X
wb = load_workbook('PitcherX_Game.xlsx')
x_data = wb["Sheet1"]
requested_pitch = sys.argv[1]

# empty arrays hold data for plotting
# hc stands for hitters count, pc is pitchers count
# right/left stands for which side the batter is on
hc_right_plate_data = []
hc_right_plate_x = []
hc_right_plate_y = []
hc_right_count = []
hc_left_plate_data = []
hc_left_plate_x = []
hc_left_plate_y = []
hc_left_count = []

pc_right_plate_data = []
pc_right_plate_x = []
pc_right_plate_y = []
pc_right_count = []
pc_left_plate_data = []
pc_left_plate_x = []
pc_left_plate_y = []
pc_left_count = []

all_right_plate_data = []
all_right_plate_x = []
all_right_plate_y = []
all_right_count = []
all_left_plate_data = []
all_left_plate_x = []
all_left_plate_y = []
all_left_count = []

#append data to the appropriate arrays created above
# pitch_data and batter_side come from the script below starting on line 74
def append_data(batter_side, pitch_data):
    if batter_side == "Right" and (pitch_data[2] == "3 - 1" or pitch_data[2] == "2 - 0"):
        hc_right_plate_data.append(pitch_data)
        hc_right_plate_x.append(pitch_data[3])
        hc_right_plate_y.append(pitch_data[4])
        hc_right_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))
    elif  batter_side == "Left" and (pitch_data[2] == "3 - 1" or pitch_data[2] == "2 - 0"):
        hc_left_plate_data.append(pitch_data)
        hc_left_plate_x.append(pitch_data[3])
        hc_left_plate_y.append(pitch_data[4])
        hc_left_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))
    elif batter_side == "Right" and (pitch_data[2] == "2 - 2" or pitch_data[2] == "1 - 2" or pitch_data == "0 - 2"):
        pc_right_plate_data.append(pitch_data)
        pc_right_plate_x.append(pitch_data[3])
        pc_right_plate_y.append(pitch_data[4])
        pc_right_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))
    elif  batter_side == "Left" and (pitch_data[2] == "2 - 2" or pitch_data[2] == "1 - 2" or pitch_data == "0 - 2"):
        pc_left_plate_data.append(pitch_data)
        pc_left_plate_x.append(pitch_data[3])
        pc_left_plate_y.append(pitch_data[4])
        pc_left_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))

    if  batter_side == "Right":
        all_right_plate_data.append(pitch_data)
        all_right_plate_x.append(pitch_data[3])
        all_right_plate_y.append(pitch_data[4])
        all_right_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))
    elif  batter_side == "Left":
        all_left_plate_data.append(pitch_data)
        all_left_plate_x.append(pitch_data[3])
        all_left_plate_y.append(pitch_data[4])
        all_left_count.append("%s, count: %s, %s, pitch: %s" % (pitch_data[0], pitch_data[2], pitch_data[5], pitch_data[1]))

# loop through all pitches and create an array of relevant data for each pitch.  Append individual
# pitch arrays to larger arrays created above
for i in range(2, 89):
    pitch = x_data.cell(row=i, column=15).value
    if pitch == requested_pitch or requested_pitch=="All":
        pitch_number = i - 1
        ab_count = "%s - %s" % (x_data.cell(row=i, column=13).value, x_data.cell(row=i, column=14).value)
        plate_loc_side = x_data.cell(row=i, column=37).value
        plate_loc_height = x_data.cell(row=i, column=36).value
        pitch_call = x_data.cell(row=i, column=17).value
        pitch_data = [pitch, pitch_number, ab_count, plate_loc_side, plate_loc_height, pitch_call]
        batter_side = x_data.cell(row=i, column=8).value
        append_data(batter_side, pitch_data)


# prep data for scatter plot
hc_right_trace = go.Scatter(
    x = hc_right_plate_x,
    y = hc_right_plate_y,
    mode = 'markers',
    text = hc_right_count
)
hc_right_data = [hc_right_trace]

hc_left_trace = go.Scatter(
    x = hc_left_plate_x,
    y = hc_left_plate_y,
    mode = 'markers',
    text = hc_left_count
)
hc_left_data = [hc_left_trace]

pc_right_trace = go.Scatter(
    x = pc_right_plate_x,
    y = pc_right_plate_y,
    mode = 'markers',
    text = pc_right_count
)
pc_right_data = [pc_right_trace]

pc_left_trace = go.Scatter(
    x = pc_left_plate_x,
    y = pc_left_plate_y,
    mode = 'markers',
    text = pc_left_count
)
pc_left_data = [pc_left_trace]

all_right_trace = go.Scatter(
    x = all_right_plate_x,
    y = all_right_plate_y,
    mode = 'markers',
    text = all_right_count
)
all_right_data = [all_right_trace]

all_left_trace = go.Scatter(
    x = all_left_plate_x,
    y = all_left_plate_y,
    mode = 'markers',
    text = all_left_count
)
all_left_data = [all_left_trace]

# set scatter axis layouts to stay uniform
layout = go.Layout(
    xaxis=dict(
        range=[-1.5,1.5]
    ),
    yaxis=dict(
        range=[0,4]
    )
)

# set figures with data and layouts
hc_right_fig = go.Figure(data=hc_right_data, layout=layout)
hc_left_fig = go.Figure(data=hc_left_data, layout=layout)
pc_right_fig = go.Figure(data=pc_right_data, layout=layout)
pc_left_fig = go.Figure(data=pc_left_data, layout=layout)
all_right_fig = go.Figure(data=all_right_data, layout=layout)
all_left_fig = go.Figure(data=all_left_data, layout=layout)

#create filenames
hc_right_file = "%s/hc-right-scatter.html" % (requested_pitch)
hc_left_file = "%s/hc-left-scatter.html" % (requested_pitch)
pc_right_file = "%s/pc-right-scatter.html" % (requested_pitch)
pc_left_file = "%s/pc-left-scatter.html" % (requested_pitch)
all_right_file = "%s/all-right-scatter.html" % (requested_pitch)
all_left_file = "%s/all-left-scatter.html" % (requested_pitch)

# generate plots!
plotly.offline.plot(hc_right_fig, filename=hc_right_file)
plotly.offline.plot(hc_left_fig, filename=hc_left_file)
plotly.offline.plot(pc_right_fig, filename=pc_right_file)
plotly.offline.plot(pc_left_fig, filename=pc_left_file)
plotly.offline.plot(all_right_fig, filename=all_right_file)
plotly.offline.plot(all_left_fig, filename=all_left_file)
