from openpyxl import load_workbook

wb = load_workbook('PitcherX_Game.xlsx')
x_data = wb["Sheet1"]

strike_count = 0
ball_count = 0
strikeouts = 0
walks = 0
balls_in_play = 0
hit_distr = {}
non_K = 0
runs = 0
batters_faced = 0

for i in range(2, 89):
    pitch_call = x_data.cell(row=i, column=17).value
    strikeout_BB = x_data.cell(row=i, column=18).value
    hit_type = x_data.cell(row=i, column=20).value
    out_on_play = x_data.cell(row=i, column=21).value
    run_on_play = x_data.cell(row=i, column=22).value
    plate_appearance = x_data.cell(row=i, column=2).value
    previous_appearance = x_data.cell(row=i-1, column=2).value
    if pitch_call == "BallCalled":
        ball_count = ball_count + 1
    elif pitch_call == "InPlay":
        balls_in_play = balls_in_play + 1
        strike_count = strike_count + 1
    else:
        strike_count = strike_count + 1
    if strikeout_BB == "Strikeout":
        strikeouts = strikeouts + 1
    elif strikeout_BB == "Walk":
        walks = walks + 1
    if hit_type != "Out" and hit_type:
        hit_distr[hit_type] = hit_distr.get(hit_type, 0) + 1
    if plate_appearance != previous_appearance:
        batters_faced = batters_faced + 1
    non_K = non_K + out_on_play
    runs = runs + run_on_play

IP = (strikeouts + non_K) / 3

print "Total Innings: %s" % (IP)
print "Total Runs: %s" % (runs)
print "Total Strikes: %s" % (strike_count)
print "Total Balls: %s" % (ball_count)
print "Total Strikeouts: %s" % (strikeouts)
print "Total Walks: %s" % (walks)
print "Total BIPs: %s" % (balls_in_play)
print "Total Batters Faced: %s" % (batters_faced)
print "Hit Distribution: %s" % (hit_distr)
