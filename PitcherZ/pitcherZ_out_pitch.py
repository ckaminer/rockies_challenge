from openpyxl import load_workbook

wb = load_workbook('PitcherZ_Game.xlsx')
z_data = wb["TMGamesEddie"]

pitch_count = {}
two_out_pitch_selection = {}
strikeout_pitch = {}
other_out_selection = {}

for i in range(2, 85):
    strikes = z_data.cell(row=i, column=14).value
    strikeout = z_data.cell(row=i, column=18).value
    other_out = z_data.cell(row=i, column=20).value
    pitch = z_data.cell(row=i, column=15).value
    pitch_count[pitch] = pitch_count.get(pitch, 0) + 1
    if strikes == 2:
        two_out_pitch_selection[pitch] = two_out_pitch_selection.get(pitch, 0) + 1
    if strikeout == "Strikeout":
        strikeout_pitch[pitch] = strikeout_pitch.get(pitch, 0) + 1
    if other_out == "Out":
        other_out_selection[pitch] = other_out_selection.get(pitch, 0) + 1

print pitch_count
print two_out_pitch_selection
print strikeout_pitch
print other_out_selection
