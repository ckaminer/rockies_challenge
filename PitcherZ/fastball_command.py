from openpyxl import load_workbook
import plotly
import plotly.graph_objs as go

# load data for pitcher Z
wb = load_workbook('PitcherZ_Game.xlsx')
z_data = wb["TMGamesEddie"]

# create empty array to eventually hold coordinates for pitch locations
right_plate_data = []
right_plate_x = []
right_plate_y = []
right_count = []
left_plate_data = []
left_plate_x = []
left_plate_y = []
left_count = []

def append_data(batter_side, pitch_data):
    if batter_side == "Right":
        right_plate_data.append(pitch_data)
        right_plate_x.append(pitch_data[3])
        right_plate_y.append(pitch_data[4])
        right_count.append(pitch_data[2])
    elif  batter_side == "Left":
        left_plate_data.append(pitch_data)
        left_plate_x.append(pitch_data[3])
        left_plate_y.append(pitch_data[4])
        left_count.append(pitch_data[2])


# loop through all pitches and create an array of relevant data for each pitch.  Append smaller arrays
# to larger array created above
for i in range(2, 85):
    pitch = z_data.cell(row=i, column=15).value
    if pitch == "Fastball":
        pitch_number = i - 1
        ab_count = "%s - %s" % (z_data.cell(row=i, column=13).value, z_data.cell(row=i, column=14).value)
        plate_loc_side = z_data.cell(row=i, column=37).value
        plate_loc_height = z_data.cell(row=i, column=36).value
        pitch_data = [pitch, pitch_number, ab_count, plate_loc_side, plate_loc_height]
        batter_side = z_data.cell(row=i, column=8).value
        append_data(batter_side, pitch_data)

# prep data for scatter plot
right_trace = go.Scatter(
    x = right_plate_x,
    y = right_plate_y,
    mode = 'markers',
    text = right_count
)
right_data = [right_trace]

left_trace = go.Scatter(
    x = left_plate_x,
    y = left_plate_y,
    mode = 'markers',
    text = left_count
)
left_data = [left_trace]

#create offline scatter plot (local HTML file)
plotly.offline.plot(right_data, filename='right-fastball-scatter')
plotly.offline.plot(left_data, filename='left-fastball-scatter')
